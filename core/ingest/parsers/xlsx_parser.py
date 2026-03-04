from __future__ import annotations
from typing import Dict, List

from core.utils.optional_deps import require


def parse_xlsx(path: str, max_rows: int = 20) -> Dict:
    ok, msg = require(
        "openpyxl",
        "pip install openpyxl",
        "XLSX ingest",
    )
    if not ok:
        return {"text": "", "error": msg}
    import openpyxl
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheets: List[str] = wb.sheetnames
        rows = []
        if sheets:
            ws = wb[sheets[0]]
            for idx, row in enumerate(ws.iter_rows(values_only=True)):
                if idx >= max_rows:
                    break
                rows.append(["" if v is None else str(v) for v in row])
        text_lines = ["\t".join(r) for r in rows if r]
        return {"text": "\n".join(text_lines), "error": None, "sheets": sheets}
    except (OSError, ValueError, TypeError, KeyError, AttributeError, RuntimeError) as e:
        return {"text": "", "error": str(e)}
